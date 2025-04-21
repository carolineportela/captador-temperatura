from openpyxl import Workbook, load_workbook
from os.path import exists
import os

def salvar_dados_excel(dados, caminho="dados.xlsx"):
    try:
        if exists(caminho):
            wb = load_workbook(caminho)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Data", "Hora", "Temperatura", "Umidade"])
    except:
        os.remove(caminho)
        wb = Workbook()
        ws = wb.active
        ws.append(["Data", "Hora", "Temperatura", "Umidade"])

    ws.append([dados["data"], dados["hora"], dados["temperatura"], dados["umidade"]])
    wb.save(caminho)
